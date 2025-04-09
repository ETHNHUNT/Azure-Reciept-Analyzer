import os
import sys
import time
import json
import logging
import pandas as pd
from typing import Dict, List, Optional, Union, Any, Tuple
import traceback
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeResult
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
from tenacity import retry, stop_after_attempt, wait_exponential
from datetime import datetime
from azure.core.exceptions import HttpResponseError
import glob
import re

from config import OUTPUT_DIR, MAX_RETRIES, POLLING_INTERVAL

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configure system encoding
sys.stdout.reconfigure(encoding='utf-8')

# Import openpyxl utilities if available
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# --- Define Budget Categories ---
BUDGET_CATEGORIES = [
    "Groceries", "Dining Out", "Transportation", "Utilities", "Housing",
    "Entertainment", "Clothing & Shopping", "Health & Wellness", "Personal Care",
    "Education", "Gifts & Donations", "Travel", "Pets", "Subscriptions",
    "Miscellaneous", "Income", "Savings/Investments", "Uncategorized"
]
# Create a comma-separated string for data validation formula
BUDGET_CATEGORIES_STR = f'"{",".join(BUDGET_CATEGORIES)}"'

# --- Define Excel Styles (if openpyxl is available) ---
if OPENPYXL_AVAILABLE:
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN_BORDER = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    # Style for alternating rows
    LIGHT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    ALT_ROW_STYLE = DifferentialStyle(fill=LIGHT_FILL)
    ALT_ROW_RULE = Rule(type="expression", dxf=ALT_ROW_STYLE, stopIfTrue=False)
    # Formula applies style to even rows (adjust row number as needed)
    ALT_ROW_RULE.formula = ["MOD(ROW(),2)=0"]


class AzureReceiptAnalyzer:
    """Azure Receipt Analyzer Class"""
    
    def __init__(self, endpoint: str, api_key: str, storage_connection_string: str = None, container_name: str = None):
        """
        Initialize the Azure Receipt Analyzer
        
        Args:
            endpoint (str): Azure Document Intelligence endpoint
            api_key (str): Azure Document Intelligence API key
            storage_connection_string (str, optional): Azure Blob Storage connection string (not used)
            container_name (str, optional): Azure Blob Storage container name (not used)
        """
        # Initialize Azure Document Intelligence Client
        self.client = DocumentIntelligenceClient(
            endpoint=endpoint, 
            credential=AzureKeyCredential(api_key)
        )
        
        # Rate limiting settings
        self.min_request_interval = 5.0  # seconds between requests
        self.last_request_time = 0

    def _wait_for_rate_limit(self):
        """Wait to avoid hitting rate limits, with longer delays for free tier"""
        elapsed = time.time() - self.last_request_time
        min_wait_time = 30.0  # 30 seconds minimum wait time for free tier
        
        if elapsed < min_wait_time:
            wait_time = min_wait_time - elapsed
            logger.info(f"⏱️ Rate limiting: waiting {wait_time:.2f} seconds before next request")
            time.sleep(wait_time)
            
        self.last_request_time = time.time()

    def _get_empty_result(self, image_id: Optional[str] = None) -> Dict:
        """Helper method to return an empty result structure with default values"""
        return {
            "image_id": image_id or "unknown",
            "extracted_text": "",
            "receipt_type": None,
            "merchant": {
                "name": "",
                "address": "",
                "phone": "",
                "tax_id": "",
                "gst_hst_number": ""
            },
            "transaction": {
                "date": "",
                "time": "",
                "total": "",
                "subtotal": "",
                "tax": "",
                "tax_details": {
                    "total_tax": "",
                    "tax_rates": [],
                    "tax_amounts": [],
                    "tax_types": [],
                    "has_hst": False
                }
            },
            "items": [],
            "totals": [],
            "payment": {
                "type": "",
                "amount": "",
                "card_number": "",
                "change_due": ""
            }
        }

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
    def analyze_receipt(self, image_url=None, image_data=None, image_id=None) -> Dict:
        """
        Analyze a receipt with manual polling, rate limiting, and enhanced logging.
        
        Args:
            image_url (str, optional): URL of the receipt image to analyze
            image_data (bytes, optional): Binary content of the receipt image
            image_id (str, optional): Image identifier for the receipt
            
        Returns:
            Dict: Receipt analysis results or an empty result structure on failure.
        """
        if not image_url and not image_data:
            logger.error("❌ Either image_url or image_data must be provided")
            return self._get_empty_result(image_id)
            
        # Set image name for logging
        if image_id:
            image_name = image_id
        elif image_url:
            image_name = os.path.basename(image_url)
        else:
            image_name = f"receipt_{int(time.time())}"
        
        logger.info(f"🚀 Starting analysis for: {image_name}")
        
        try:
            self._wait_for_rate_limit()
            logger.info(f"✅ Submitting request for: {image_name}...")
            
            # Start the document analysis operation
            request = AnalyzeDocumentRequest(url_source=image_url) if image_url else image_data
            poller = self.client.begin_analyze_document("prebuilt-receipt", request)
            
            logger.info(f"⏳ Polling started for operation ID: {poller.details.get('id', 'N/A')} for {image_name}")

            # Implement manual polling with timeout
            start_time = time.time()
            max_wait_time = 60  # Increased to 60 seconds for free tier
            poll_interval = 5   # Start with 5 seconds for free tier
            
            while True:
                # Check operation status
                status = poller.status()
                logger.debug(f"Polling status for {image_name}: {status}")

                if status == "succeeded":
                    logger.info(f"✅ Operation succeeded for {image_name} after {time.time() - start_time:.2f} seconds.")
                    break
                elif status == "failed":
                    logger.error(f"❌ Operation failed for {image_name} after {time.time() - start_time:.2f} seconds.")
                    # Try to get error details if available
                    error_info = "No specific error details available from poller."
                    if hasattr(poller, 'details') and poller.details:
                       error_info = poller.details.get('error', error_info)
                    logger.error(f"Error details: {error_info}")
                    return self._get_empty_result(image_name) # Return empty on failure
                
                # Check for timeout
                if (time.time() - start_time) > max_wait_time:
                    logger.error(f"❌ Operation timed out for {image_name} after {max_wait_time} seconds.")
                    poller.cancel() # Attempt to cancel the operation
                    logger.warning(f"Attempted to cancel operation for {image_name}")
                    return self._get_empty_result(image_name) # Return empty on timeout

                # Wait before next poll
                time.sleep(poll_interval)
                # Exponential backoff for polling interval (optional, simple fixed interval here)
                # poll_interval = min(poll_interval * 1.5, 15) 
            
            # Retrieve the result
            logger.info(f"🔄 Retrieving results for {image_name}...")
            try:
                receipt_result: AnalyzeResult = poller.result()
                # Log the raw structure immediately after getting it
                logger.debug(f"Raw AnalyzeResult structure received for {image_name}: {receipt_result}") 
                
                if not receipt_result or not receipt_result.documents:
                    logger.warning(f"⚠️ Result retrieved, but no documents found in the result for {image_name}.")
                    return self._get_empty_result(image_name)
                    
                logger.info(f"✅ Successfully retrieved results for {image_name}. Processing...")
                return self._process_receipt_result(receipt_result, image_id=image_name) # Pass image_id for context

            except Exception as e:
                 logger.error(f"❌ Error retrieving or processing poller result for {image_name}: {str(e)}\n{traceback.format_exc()}")
                 return self._get_empty_result(image_name)

        except HttpResponseError as e:
            logger.error(f"❌ Azure HTTP Error during analysis for {image_name}: {e.status_code} - {e.message}\n{traceback.format_exc()}")
            return self._get_empty_result(image_name)
        except Exception as e:
            logger.error(f"❌ Unexpected error during analysis for {image_name}: {str(e)}\n{traceback.format_exc()}")
            return self._get_empty_result(image_name)

    def _add_to_retry_queue(self, image_url: str = "", image_id: str = ""):
        """Add a failed analysis to the retry queue"""
        # Get or create output directory
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        retry_file = os.path.join(OUTPUT_DIR, "receipt_retry_queue.json")
        retry_queue = []
        
        # Load existing queue if it exists
        if os.path.exists(retry_file):
            try:
                with open(retry_file, 'r') as f:
                    retry_queue = json.load(f)
            except Exception as e:
                logger.error(f"Error loading retry queue: {str(e)}")
                retry_queue = []
        
        # Add new entry
        retry_queue.append({
            "image_url": image_url,
            "image_id": image_id,
            "timestamp": datetime.now().isoformat()
        })
        
        # Save updated queue
        try:
            with open(retry_file, 'w') as f:
                json.dump(retry_queue, f, indent=2)
        except Exception as e:
            logger.error(f"Error saving retry queue: {str(e)}")

    def _clean_quantity(self, value: Any) -> str:
        """Clean and convert quantity values to string, handling both integers and decimals."""
        if value is None or value == "":
            return "1"
        
        try:
            # If already a number, convert to float first
            if isinstance(value, (int, float)):
                qty = float(value)
            else:
                # Remove any non-numeric characters except decimal point
                clean_value = ''.join(c for c in str(value) if c.isdigit() or c in ['.'])
                qty = float(clean_value) if clean_value else 1.0
            
            # Format the quantity appropriately
            if qty == int(qty):  # If it's a whole number
                return str(int(qty))
            return str(qty)  # Keep decimals for fractional quantities
        except (ValueError, TypeError):
            logger.warning(f"Could not clean quantity value: {value}, using default of 1")
            return "1"

    def _process_receipt_result(self, receipts: AnalyzeResult, image_id: Optional[str] = None) -> Dict:
        """Process the receipt analysis result"""
        try:
            # Create the base result structure, using the provided image_id
            result = self._get_empty_result(image_id=image_id)
            
            # Log the full raw result structure
            logger.debug(f"Processing AnalyzeResult for {image_id}: {receipts}")
            
            # Check receipts object itself before accessing documents
            if not receipts:
                 logger.error(f"❌ Received None or empty AnalyzeResult object for {image_id}")
                 return result
                 
            if not receipts.documents:
                logger.warning(f"❌ No documents found in the AnalyzeResult for {image_id}")
                return result

            doc = receipts.documents[0]
            logger.info(f"✅ Processing receipt of type: {doc.doc_type}")
            
            # Set basic information
            result["image_id"] = image_id # Use the passed image_id
            result["extracted_text"] = receipts.content if receipts.content else ''
            result["receipt_type"] = doc.doc_type if doc.doc_type else None
            
            # Log document structure for debugging
            logger.debug(f"Document fields: {doc.fields.keys() if hasattr(doc, 'fields') else 'No fields'}")
            
            # If there are no fields, return the basic result
            if not hasattr(doc, 'fields') or not doc.fields:
                logger.warning("❌ Document has no fields to process")
                return result
            
            # Initialize tax variables
            total_tax = None # Use None initially
            tax_rates = []
            tax_amounts = []
            tax_types = []
            subtotal = None # Use None initially
            total = None    # Use None initially
            has_hst = False

            # --- Pre-process to find HST number early ---
            # Look for HST/GST number in MerchantTaxId or potentially specific regex in content
            hst_number_found = False
            if 'MerchantTaxId' in doc.fields:
                 tax_id_field = doc.fields['MerchantTaxId']
                 tax_id_value = self._extract_field_value(tax_id_field)
                 if tax_id_value:
                     tax_id_str = str(tax_id_value)
                     # Regex to find patterns like HST# XXXXXXXXX RTXXXX or GST# XXXXXXXXX
                     match = re.search(r'(?:HST|GST)\s*#?\s*(\d{9})\s*(?:RT\d{4})?', tax_id_str, re.IGNORECASE)
                     if match:
                         result["merchant"]["gst_hst_number"] = f"{match.group(0)}" # Store the full matched string
                         has_hst = True
                         hst_number_found = True
                         logger.info(f"Found GST/HST number via MerchantTaxId: {result['merchant']['gst_hst_number']}")
                     else:
                         # If it doesn't match HST/GST pattern, store as generic tax ID
                         result["merchant"]["tax_id"] = tax_id_str

            # --- Process main fields ---
            for field_name, field in doc.fields.items():
                try:
                    value = self._extract_field_value(field)
                    logger.debug(f"Processing field {field_name}: {value}")
                    
                    # Handle merchant fields (name, address, phone already done partially)
                    if field_name == 'MerchantName':
                         result["merchant"]["name"] = str(value or '')
                         logger.info(f"Found merchant name: {result['merchant']['name']}")
                    elif field_name == 'MerchantAddress':
                         result["merchant"]["address"] = str(value or '')
                    elif field_name == 'MerchantPhoneNumber':
                         result["merchant"]["phone"] = str(value or '')
                    # Skip MerchantTaxId here as it was pre-processed

                    # Handle transaction fields
                    elif field_name == 'TransactionDate':
                        result["transaction"]["date"] = str(value or '')
                    elif field_name == 'TransactionTime':
                        result["transaction"]["time"] = str(value or '')
                    elif field_name == 'Total':
                        cleaned_total = self._clean_currency(value)
                        if cleaned_total is not None:
                            total = cleaned_total
                            result["transaction"]["total"] = str(total)
                            logger.info(f"Found total: {total}")
                    elif field_name == 'Subtotal':
                        cleaned_subtotal = self._clean_currency(value)
                        if cleaned_subtotal is not None:
                            subtotal = cleaned_subtotal
                            result["transaction"]["subtotal"] = str(subtotal)
                            logger.info(f"Found subtotal: {subtotal}")
                    elif field_name == 'TotalTax': # Capture top-level TotalTax if available
                         cleaned_tax = self._clean_currency(value)
                         if cleaned_tax is not None:
                              total_tax = cleaned_tax
                              logger.info(f"Found top-level TotalTax: {total_tax}")
                              # If we found HST number earlier, assume this tax is related
                              if hst_number_found and not tax_amounts:
                                   tax_amounts.append(total_tax)
                                   tax_types.append("HST/GST") # Generic type

                    # Handle detailed tax fields (preferred if available)
                    elif field_name == 'TaxDetails' and hasattr(field, 'value_array') and field.value_array:
                        logger.info("Processing TaxDetails array...")
                        tax_amounts = [] # Reset if we get details
                        tax_types = []   # Reset if we get details
                        tax_rates = []   # Reset if we get details
                        current_total_tax = 0.0 # Recalculate from details
                        for tax_item in field.value_array:
                            if hasattr(tax_item, 'value_object'):
                                tax_obj = tax_item.value_object
                                tax_type = ""
                                rate = None
                                amount = None

                                if 'Description' in tax_obj:
                                    tax_type = str(self._extract_field_value(tax_obj['Description']) or '')
                                if 'Rate' in tax_obj:
                                    rate_val = self._extract_field_value(tax_obj['Rate'])
                                    if rate_val:
                                         try: # Try converting rate percentage/number
                                              if isinstance(rate_val, str) and '%' in rate_val: rate = float(rate_val.replace('%','')) / 100
                                              elif isinstance(rate_val, (int, float)): rate = float(rate_val)/100 if rate_val > 1 else float(rate_val)
                                         except: rate = None # Ignore invalid rates
                                if 'Amount' in tax_obj:
                                    amount = self._clean_currency(self._extract_field_value(tax_obj['Amount']))

                                if amount is not None:
                                     tax_amounts.append(amount)
                                     tax_types.append(tax_type)
                                     tax_rates.append(rate) # Append rate (even if None)
                                     current_total_tax += amount
                                     logger.info(f"Found tax detail: Type='{tax_type}', Rate={rate}, Amount={amount}")
                                     if any(t in tax_type.upper() for t in ['HST', 'GST', 'PST']):
                                         has_hst = True # Set based on specific types found

                        total_tax = current_total_tax # Update total tax from details
                        logger.info(f"Total tax recalculated from details: {total_tax}")
                    
                    # Handle Payment Details
                    elif field_name == 'PaymentDetails' and hasattr(field, 'value_array'):
                         logger.info("Processing PaymentDetails...")
                         for payment_item in field.value_array:
                              if hasattr(payment_item, 'value_object'):
                                   payment_obj = payment_item.value_object
                                   # Heuristic: Assume the first payment detail is the primary one for now
                                   if not result["payment"]["type"] and 'PaymentType' in payment_obj:
                                        result["payment"]["type"] = str(self._extract_field_value(payment_obj['PaymentType']) or '')
                                   if not result["payment"]["amount"] and 'Amount' in payment_obj:
                                        pay_amount = self._clean_currency(self._extract_field_value(payment_obj['Amount']))
                                        if pay_amount is not None: result["payment"]["amount"] = str(pay_amount)
                                   if not result["payment"]["card_number"] and 'CreditCardLast4Digits' in payment_obj:
                                        result["payment"]["card_number"] = "****" + str(self._extract_field_value(payment_obj['CreditCardLast4Digits']) or '')
                                   logger.debug(f"Processed payment detail object: {payment_obj}")

                    elif field_name == 'AmountTendered': # Handle separate AmountTendered if PaymentDetails doesn't have it
                         if not result["payment"]["amount"]:
                              pay_amount = self._clean_currency(value)
                              if pay_amount is not None: result["payment"]["amount"] = str(pay_amount)
                    elif field_name == 'ChangeDue':
                         change = self._clean_currency(value)
                         if change is not None: result["payment"]["change_due"] = str(change)
                    elif field_name == 'PaymentType': # Handle separate PaymentType
                         if not result["payment"]["type"]: result["payment"]["type"] = str(value or '')


                    # Handle items with enhanced tax status and savings
                    elif field_name == 'Items' and hasattr(field, 'value_array'):
                        logger.info("Processing items...")
                        result["items"] = [] # Ensure items list is reset
                        for item in field.value_array:
                            item_dict = {
                                "description": "", "quantity": "1", "price": "", "total": "",
                                "tax_status": "UNKNOWN", "tax_amount": "", "final_price": "",
                                "discount": "", "savings": ""
                            }
                            
                            if hasattr(item, 'value_object'):
                                item_obj = item.value_object
                                logger.debug(f"Processing item object: {item_obj}")
                                
                                # Extract standard item fields
                                if 'Description' in item_obj: item_dict["description"] = str(self._extract_field_value(item_obj['Description']) or '')
                                if 'Quantity' in item_obj: item_dict["quantity"] = self._clean_quantity(self._extract_field_value(item_obj['Quantity']))
                                if 'Price' in item_obj: # Unit Price
                                     price = self._clean_currency(self._extract_field_value(item_obj['Price']))
                                     if price is not None: item_dict["price"] = str(price)
                                if 'TotalPrice' in item_obj: # Line Total
                                     total_price = self._clean_currency(self._extract_field_value(item_obj['TotalPrice']))
                                     if total_price is not None: item_dict["total"] = str(total_price)
                                
                                # --- Infer Price/Total if one is missing ---
                                if item_dict["total"] and not item_dict["price"] and item_dict["quantity"]:
                                     try:
                                          qty_f = float(item_dict["quantity"])
                                          total_f = float(item_dict["total"])
                                          if qty_f > 0: item_dict["price"] = str(round(total_f / qty_f, 2))
                                     except: pass # Ignore errors
                                elif item_dict["price"] and not item_dict["total"] and item_dict["quantity"]:
                                      try:
                                          qty_f = float(item_dict["quantity"])
                                          price_f = float(item_dict["price"])
                                          item_dict["total"] = str(round(price_f * qty_f, 2))
                                      except: pass # Ignore errors

                                # --- Handle Tax Status ---
                                # Default based on overall HST status, then check keywords
                                item_tax_status = "TAXABLE" if has_hst else "UNKNOWN"
                                item_desc_upper = item_dict["description"].upper()
                                if 'ZERO' in item_desc_upper or '0%' in item_desc_upper:
                                    item_tax_status = "ZERO-RATED"
                                elif 'EXEMPT' in item_desc_upper:
                                    item_tax_status = "EXEMPT"
                                # TODO: Add logic for specific codes like 'C', 'HC' if needed and mappings are known
                                item_dict["tax_status"] = item_tax_status

                                # --- Calculate Item Tax Amount (if possible) ---
                                if item_tax_status == "TAXABLE" and has_hst and item_dict["total"]:
                                     # Assuming primary HST rate (e.g., 13% in ON) - This is an approximation!
                                     # A more accurate method would require knowing which tax rate applies.
                                     primary_rate = tax_rates[0] if tax_rates else 0.13 # Use first rate or default
                                     try:
                                          line_total_f = float(item_dict["total"])
                                          item_tax_amount = round(line_total_f * primary_rate, 2)
                                          item_dict["tax_amount"] = str(item_tax_amount)
                                          item_dict["final_price"] = str(round(line_total_f + item_tax_amount, 2))
                                     except:
                                          item_dict["tax_amount"] = "" # Clear if calculation fails
                                          item_dict["final_price"] = item_dict["total"] # Fallback
                                elif item_dict["total"]: # If not taxable or no HST, final price is just total
                                     item_dict["final_price"] = item_dict["total"]

                                # --- Extract Discount/Savings ---
                                if 'Discount' in item_obj: # Check specific discount field
                                     discount = self._clean_currency(self._extract_field_value(item_obj['Discount']))
                                     if discount is not None: item_dict["discount"] = str(discount)
                                # Crude check for "YOU SAVED" in description (less reliable)
                                savings_match = re.search(r'YOU SAVED \$?([\d.]+)', item_dict["description"], re.IGNORECASE)
                                if savings_match:
                                     savings = self._clean_currency(savings_match.group(1))
                                     if savings is not None: item_dict["savings"] = str(savings)
                            
                            if item_dict["description"]:  # Only add items with descriptions
                                result["items"].append(item_dict)
                                logger.info(f"Added item: {item_dict['description']}")
                except Exception as e:
                    logger.error(f"Error processing field {field_name}: {str(e)}\n{traceback.format_exc()}")
                    continue

            # --- Final Tax Details Population ---
            result["transaction"]["tax"] = str(total_tax) if total_tax is not None else ""
            result["transaction"]["tax_details"].update({
                "total_tax": str(total_tax) if total_tax is not None else "",
                "tax_rates": [r for r in tax_rates if r is not None], # Store valid rates
                "tax_amounts": [float(amt) for amt in tax_amounts if amt is not None], # Store valid amounts
                "tax_types": [t for t in tax_types if t], # Store non-empty types
                "has_hst": has_hst
            })

            # Log final result summary
            logger.info(f"Receipt processing complete for {image_id}. Found {len(result['items'])} items.")
            logger.info(f"Total: {result['transaction']['total']}, Subtotal: {result['transaction']['subtotal']}, Tax: {result['transaction']['tax']}")
            logger.info(f"Payment: Type='{result['payment']['type']}', Amount='{result['payment']['amount']}', Change='{result['payment']['change_due']}'")
            
            return result
        
        except Exception as e:
            logger.error(f"Critical error processing receipt result for {image_id}: {str(e)}\n{traceback.format_exc()}")
            # Ensure we return the basic empty structure even on major failure
            return self._get_empty_result(image_id=image_id)

    def _extract_field_value(self, field: Any) -> Optional[Union[str, float, int, List, Dict]]:
        """Extracts the value from various Azure Document Intelligence field types."""
        if not field:
            return None

        field_type = getattr(field, 'kind', None) or getattr(field, 'type', None) # Compatibility for different SDK versions
        logger.debug(f"Extracting value from field type: {field_type}, Content: {getattr(field, 'content', 'N/A')}")

        if hasattr(field, 'content') and field.content is not None:
             # Prioritize content if available and non-empty
             return field.content
        elif field_type == 'string' and hasattr(field, 'value_string'):
             return field.value_string
        elif field_type == 'number' and hasattr(field, 'value_number'):
             return field.value_number
        elif field_type == 'integer' and hasattr(field, 'value_integer'):
            return field.value_integer
        elif field_type == 'date' and hasattr(field, 'value_date'):
             return str(field.value_date) # Convert date to string
        elif field_type == 'phoneNumber' and hasattr(field, 'value_phone_number'):
             return field.value_phone_number
        elif field_type == 'countryRegion' and hasattr(field, 'value_country_region'):
            return field.value_country_region
        elif field_type == 'currency' and hasattr(field, 'value_currency'):
             # Format currency object if needed, or return amount
             amount = getattr(field.value_currency, 'amount', None)
             symbol = getattr(field.value_currency, 'currency_symbol', '')
             return amount # Just return the amount for now
        elif field_type == 'address' and hasattr(field, 'value_address'):
            # Return the content/raw string for address for simplicity
            return getattr(field, 'content', None)
        # Add more specific type handling if needed (e.g., value_time, value_array, value_object)
        # elif field_type == 'array' and hasattr(field, 'value_array'):
        #     # Handle arrays if necessary (often handled directly in _process_receipt_result)
        #     return field.value_array 
        # elif field_type == 'object' and hasattr(field, 'value_object'):
        #      # Handle objects if necessary (often handled directly in _process_receipt_result)
        #     return field.value_object

        # Fallback if no specific value_* attribute found or content was None
        logger.warning(f"Could not extract specific value from field type {field_type}. Field details: {field}")
        return None # Explicitly return None if no value found

    def _clean_currency(self, value: Optional[Union[str, float, int]]) -> Optional[float]:
        """Clean and convert currency values to float. Returns None if invalid."""
        if value is None or value == "":
            return None # Use None for missing values for better aggregation

        try:
            if isinstance(value, (int, float)):
                return float(value)

            # Remove common currency symbols, commas, spaces, and letters (like T/H indicators)
            clean_value = re.sub(r"[$,\sA-Za-z]", "", str(value))

            if not clean_value:
                 return None

            # Allow negative numbers (e.g., for returns or discounts)
            if clean_value.count('-') > 1: # Invalid negative
                 return None
            # Ensure '-' is only at the start
            if '-' in clean_value and not clean_value.startswith('-'):
                 return None

            return float(clean_value)

        except (ValueError, TypeError):
            logger.warning(f"Could not clean currency value: {value}")
            return None # Return None on error

    def _categorize_item(self, description: str) -> str:
        """Categorize an item based on its description."""
        description = description.lower() if description else ""
        
        # Define category mappings with common items
        category_mappings = {
            "groceries": ["milk", "bread", "cheese", "yogurt", "fruit", "vegetable", "meat", "chicken", "beef", "pork", "fish", "egg", "cereal", "pasta", "rice", "flour", "sugar", "coffee", "tea", "juice", "water", "snack", "chip", "cookie", "cracker", "nut", "bean", "grain", "produce", "dairy", "bakery", "deli", "frozen", "canned", "pantry", "spice", "oil", "vinegar", "sauce", "condiment", "paneer", "naan", "tandoori"],
            "household": ["paper", "toilet", "tissue", "detergent", "soap", "cleaner", "cleaning", "laundry", "dish", "towel", "trash", "garbage", "bag", "battery", "light bulb", "candle", "filter", "storage", "container", "foil", "wrap", "ziploc", "vacuum", "broom", "mop", "sponge", "glove", "supply"],
            "personal_care": ["shampoo", "conditioner", "body wash", "soap", "lotion", "cream", "deodorant", "toothpaste", "toothbrush", "floss", "mouthwash", "razor", "shaving", "makeup", "cosmetic", "cotton", "sanitizer", "medicine", "vitamin", "supplement", "bandage", "sunscreen", "insect repellent", "nail", "hair"],
            "dining": ["restaurant", "cafe", "coffee shop", "bar", "pub", "fast food", "takeout", "delivery", "pizza", "burger", "sandwich", "salad", "breakfast", "lunch", "dinner", "meal", "drink", "beverage", "alcohol", "beer", "wine", "liquor", "cocktail", "appetizer", "dessert", "tip", "service charge"],
            "transportation": ["gas", "fuel", "parking", "toll", "fare", "ticket", "uber", "lyft", "taxi", "cab", "bus", "train", "subway", "rental", "car wash", "oil change", "maintenance", "repair", "tire", "battery", "insurance", "registration", "license", "dmv"],
            "entertainment": ["movie", "theater", "concert", "show", "event", "ticket", "game", "book", "magazine", "music", "video", "streaming", "subscription", "toy", "game", "hobby", "craft", "sport", "fitness", "gym", "class", "lesson", "tour", "park", "museum", "attraction"]
        }
        
        # Check for matches in each category
        for category, keywords in category_mappings.items():
            for keyword in keywords:
                if keyword in description:
                    # Convert to display format (capitalize words)
                    return " ".join(word.capitalize() for word in category.split("_"))
        
        # For reusable bags specifically
        if "bag" in description and "reusable" in description:
            return "Household"
            
        # Default category if no match found
        return "Groceries"

    def process_receipts(self, file_paths: List[str], max_receipts: Optional[int] = 10) -> List[Dict]:
        """
        Process receipts from local files
        
        Args:
            file_paths (List[str]): List of paths to local receipt files
            max_receipts (Optional[int]): Maximum number of receipts to process. Default is 10.
            
        Returns:
            List[Dict]: List of receipt analysis results
        """
        # Get or create output directory
        output_dir = os.environ.get('OUTPUT_DIR', 'output')
        os.makedirs(output_dir, exist_ok=True)
        
        print("\n===== PROCESSING RECEIPT FILES =====")
        print(f"Max receipts to process: {max_receipts}")
        print(f"Output directory: {output_dir}")
        print("====================================\n")

        file_paths_to_process = file_paths[:max_receipts]
        all_results = []
        
        for i, file_path in enumerate(file_paths_to_process):
             try:
                 print(f"\n🔍 Processing receipt {i+1}/{len(file_paths_to_process)}: {os.path.basename(file_path)}")
                 result = self.analyze_local_receipt(file_path)
                 all_results.append(result)
                 if result.get("merchant", {}).get("name") or result.get("transaction", {}).get("total"):
                      print(f"✅ Extracted data from {os.path.basename(file_path)}")
                 else:
                      print(f"⚠️ Limited/no data extracted from {os.path.basename(file_path)}")
             except Exception as e:
                 logger.error(f"Error processing {file_path}: {e}", exc_info=True)
                 print(f"❌ Error processing {os.path.basename(file_path)}: {e}")
                 all_results.append(self._get_empty_result(image_id=os.path.basename(file_path)))

             # Optional delay (consider removing if not needed for API limits)
             # if i < len(file_paths_to_process) - 1:
             #     time.sleep(config.POLLING_INTERVAL)


        # --- Save results ---
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Save Raw JSON (always useful)
        json_file = os.path.join(output_dir, f"receipt_analysis_{timestamp}_raw.json")
        try:
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(all_results, f, indent=2)
            print(f"\n✅ Raw JSON results saved to {json_file}")
        except Exception as e:
            logger.error(f"Failed to save raw JSON: {e}")
            print(f"❌ Failed to save raw JSON results: {e}")

        # Save Enhanced Excel - construct the full path instead of passing separate arguments
        excel_file_path = os.path.join(output_dir, f"receipt_analysis_{timestamp}.xlsx")
        self._save_results_excel(all_results, excel_file_path)  # Correct call with just 2 arguments

        # --- Final Summary ---
        successful = len([r for r in all_results if r.get("merchant", {}).get("name") or r.get("transaction", {}).get("total")])
        print(f"\n📊 Analysis complete: {successful}/{len(all_results)} receipts processed.")
        if successful == 0 and all_results:
             print("\n⚠️ No receipts were successfully analyzed in detail.")
             print("   Check the raw JSON file and logs for potential errors.")
        elif successful > 0:
             print("\n💡 TIP: Open the Excel file to categorize expenses and analyze your spending!")
        print("\n=====================================")

        return all_results

    def generate_summary_report(self, results: List[Dict]) -> Dict:
        """
        Generate a summary report from the receipt analysis results
        
        Args:
            results (List[Dict]): List of receipt analysis results
            
        Returns:
            Dict: Summary report
        """
        if not results:
            return {
                "total_spending": 0, "receipt_count": 0, "merchant_count": 0,
                "item_count": 0, "merchants": [], "categories": {},
                "date_range": {"start": None, "end": None}
            }
            
        total_spending = 0
        merchants = set()
        item_count = 0
        all_dates = []
        categories = {}
        
        for receipt in results:
            total = self._clean_currency(receipt.get("transaction", {}).get("total"))
            if total is not None:
                 total_spending += total

            merchant_name = receipt.get("merchant", {}).get("name")
            if merchant_name:
                merchants.add(merchant_name)

            date_str = receipt.get("transaction", {}).get("date")
            if date_str:
                 try:
                     # Attempt to parse various date formats
                     dt = pd.to_datetime(date_str).date()
                     all_dates.append(dt)
                 except Exception:
                      logger.warning(f"Could not parse date: {date_str}")


            for item in receipt.get("items", []):
                item_count += 1
                category = self._categorize_item(item.get("description", ""))
                item_total = self._clean_currency(item.get("total")) # Use pre-tax total for category spending
                if item_total is not None:
                     categories[category] = categories.get(category, 0) + item_total

        date_range = {"start": min(all_dates).isoformat() if all_dates else None,
                      "end": max(all_dates).isoformat() if all_dates else None}

        return {
            "total_spending": round(total_spending, 2),
            "receipt_count": len(results),
            "merchant_count": len(merchants),
            "item_count": item_count,
            "merchants": sorted(list(merchants)),
            "categories": {k: round(v, 2) for k, v in sorted(categories.items(), key=lambda item: item[1], reverse=True)},
            "date_range": date_range
        }

    def analyze_local_receipt(self, file_path: str) -> Dict:
        """
        Analyze a receipt from a local file
        
        Args:
            file_path (str): Path to the local receipt file
            
        Returns:
            Dict: Receipt analysis results
        """
        try:
            # Check if file exists
            if not os.path.exists(file_path):
                logger.error(f"File does not exist: {file_path}")
                return self._get_empty_result(image_id=os.path.basename(file_path))
                
            # Check file extension
            file_ext = os.path.splitext(file_path)[1].lower()
            supported_formats = ['.jpg', '.jpeg', '.png', '.pdf', '.tif', '.tiff', '.bmp']
            
            if file_ext not in supported_formats:
                logger.error(f"Unsupported file format: {file_ext}. Supported formats: {', '.join(supported_formats)}")
                return self._get_empty_result(image_id=os.path.basename(file_path))
            
            # Read the image file
            with open(file_path, 'rb') as image_file:
                image_data = image_file.read()
            
            # Get file name as image_id
            image_id = os.path.basename(file_path)
            
            # Process using the analyze_receipt method with binary data
            result = self.analyze_receipt(image_data=image_data, image_id=image_id)
            return result
            
        except Exception as e:
            logger.error(f"Error analyzing local receipt {file_path}: {str(e)}")
            return self._get_empty_result(image_id=os.path.basename(file_path))

    def _save_results_excel(self, results, output_file_path):
        """Save results as an ultra-minimal Excel file to prevent corruption."""
        logger.info(f"Attempting to save ultra-minimal Excel to: {output_file_path}")
        try:
            # Only prepare item data
            all_items = []
            
            # Ontario HST rate
            HST_RATE = 0.13  # 13% HST in Ontario
            
            for result in results:
                receipt_id = result.get("image_id", "unknown")
                date = result.get("transaction", {}).get("date", "")
                merchant = result.get("merchant", {}).get("name", "")
                
                # Check for Ontario indicators
                is_ontario = False
                address = result.get("merchant", {}).get("address", "").upper()
                if "ON" in address or "ONTARIO" in address:
                    is_ontario = True
                
                # Get receipt totals
                subtotal = self._clean_currency(result.get("transaction", {}).get("subtotal", "0"))
                total_tax = self._clean_currency(result.get("transaction", {}).get("tax", "0"))
                
                # Process items
                for item in result.get("items", []):
                    description = item.get("description", "")
                    
                    # Convert numeric values to proper float type
                    try:
                        quantity = float(item.get("quantity", 1))
                        # If quantity is a whole number, store as integer
                        if quantity == int(quantity):
                            quantity = int(quantity)
                    except (ValueError, TypeError):
                        quantity = 1
                        
                    try:
                        price = float(item.get("price", 0))
                    except (ValueError, TypeError):
                        price = 0.0
                        
                    try:
                        total = float(item.get("total", 0))
                    except (ValueError, TypeError):
                        total = 0.0
                    
                    # Determine tax status based on content
                    raw_items = str(result.get("raw_data", ""))
                    tax_status = "EXEMPT"  # Default in Ontario (most groceries are tax-exempt)
                    
                    try:
                        tax_amount = 0.0  # Initialize as float
                        
                        # If item includes HC designation or is the "Bag Reusable" which is taxable
                        if "HC" in raw_items and description and (description in raw_items[:raw_items.find(description)+100]):
                            tax_status = "TAXABLE"
                            if is_ontario:
                                tax_amount = round(float(total) * HST_RATE, 2)
                        
                        # Special case for reusable bags which are usually taxable
                        if "bag" in description.lower() and "reusable" in description.lower():
                            tax_status = "TAXABLE"
                            if is_ontario:
                                tax_amount = round(float(total) * HST_RATE, 2)
                    except (ValueError, TypeError):
                        tax_amount = 0.0
                    
                    # Calculate final price (total + tax) as float
                    try:
                        final_price = float(total)
                        if tax_amount > 0:
                            final_price = round(float(total) + tax_amount, 2)
                    except (ValueError, TypeError):
                        final_price = 0.0
                    
                    # Auto-categorize the item based on description
                    category = self._categorize_item(description)
                    
                    # Convert savings to float if possible
                    try:
                        savings = float(item.get("savings", 0)) if item.get("savings") else None
                    except (ValueError, TypeError):
                        savings = None
                    
                    # Build item data for Excel
                    item_data = {
                        "Receipt ID": receipt_id,
                        "Date": date,
                        "Merchant": merchant,
                        "Description": description,
                        "Quantity": quantity,  # Stored as float or int
                        "Unit Price": price,   # Stored as float
                        "Total": total,        # Stored as float
                        "Tax Status": tax_status,
                        "Tax Amount": tax_amount,  # Stored as float
                        "Final Price": final_price,  # Stored as float
                        "Category": category,
                        "Savings": savings,    # Stored as float if available
                        "Notes": ""
                    }
                    all_items.append(item_data)
            
            # Create a new workbook
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Receipt Items"
            
            # Define headers and write them
            headers = [
                "Receipt ID", "Date", "Merchant", "Description", "Quantity", 
                "Unit Price", "Total", "Tax Status", "Tax Amount", "Final Price", 
                "Category", "Savings", "Notes"
            ]
            
            # Write headers with minimal styling
            for col_idx, header in enumerate(headers, 1):
                # Convert header to safe string
                safe_header = str(header).strip()
                cell = ws.cell(row=1, column=col_idx, value=safe_header)
                # Minimal header styling
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Define number format mappings
            number_formats = {
                "Quantity": "#,##0.00",  # Up to 2 decimal places for quantity
                "Unit Price": "$#,##0.00",  # Currency with 2 decimal places
                "Total": "$#,##0.00",      # Currency with 2 decimal places
                "Tax Amount": "$#,##0.00",  # Currency with 2 decimal places
                "Final Price": "$#,##0.00", # Currency with 2 decimal places
                "Savings": "$#,##0.00"      # Currency with 2 decimal places
            }
            
            # Write data with proper numeric formatting
            for row_idx, item in enumerate(all_items, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = item.get(header, "")
                    
                    # Skip empty values
                    if value is None or value == "":
                        ws.cell(row=row_idx, column=col_idx, value=None)
                        continue
                    
                    # Write cell value - directly write numbers as numbers
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply appropriate formatting based on column type
                    if header in number_formats and isinstance(value, (int, float)):
                        cell.number_format = number_formats[header]
            
            # Define a comprehensive list of categories
            categories = [
                "Groceries", 
                "Household",
                "Personal Care",
                "Dining",
                "Transportation",
                "Entertainment",
                "Education",
                "Health",
                "Clothing",
                "Electronics",
                "Gifts",
                "Other"
            ]
            
            # Find the Category column index
            category_col = None
            for idx, header in enumerate(headers, 1):
                if header == "Category":
                    category_col = idx
                    break
            
            # Create and apply dropdown for Category column
            if category_col:
                try:
                    # Create data validation
                    dv = DataValidation(
                        type="list", 
                        formula1=f'"{",".join(categories)}"',
                        allow_blank=True
                    )
                    
                    # Add the validation to worksheet
                    ws.add_data_validation(dv)
                    
                    # Apply validation to all cells in the Category column (skip header)
                    for row in range(2, len(all_items) + 2):
                        cell_location = f"{get_column_letter(category_col)}{row}"
                        dv.add(cell_location)
                        
                    logger.info(f"Successfully added category dropdown to column {get_column_letter(category_col)}")
                except Exception as e:
                    logger.warning(f"Could not add category dropdown validation: {e}")
            
            # Auto-adjust column widths for better readability
            for col_idx, header in enumerate(headers, 1):
                # Calculate appropriate width based on content
                max_length = len(str(header))
                
                # Check content length in this column (first 100 rows max)
                for row_idx in range(2, min(len(all_items) + 2, 100)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, min(len(str(cell_value)), 50))
                
                # Set column width with padding
                column = ws.column_dimensions[get_column_letter(col_idx)]
                column.width = max_length + 4  # Add padding
            
            # Save workbook
            try:
                wb.save(output_file_path)
                print(f"✅ Excel file with numeric data saved to {output_file_path}")
                logger.info("Successfully saved Excel file with proper numeric formatting")
                return output_file_path
            except Exception as e:
                logger.error(f"Excel save error: {e}")
                # Fall back to CSV
                csv_path = output_file_path.replace(".xlsx", ".csv")
                self._save_results_csv(results, csv_path)
                return csv_path
                
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            print(f"❌ Error creating Excel file: {e}")
            # Fall back to CSV
            csv_path = output_file_path.replace(".xlsx", ".csv")
            self._save_results_csv(results, csv_path)
            return csv_path

    def _save_results_csv(self, results, output_file_path):
        """Save results as CSV file (fallback if Excel fails)."""
        logger.info(f"Saving results as CSV to {output_file_path}")
        try:
            # Extract items from all receipts
            all_items = []
            for result in results:
                receipt_id = result.get("image_id", "unknown")
                date = result.get("transaction", {}).get("date", "")
                merchant = result.get("merchant", {}).get("name", "")
                
                # Process items
                for item in result.get("items", []):
                    # Convert numeric values to proper types
                    try:
                        quantity = float(item.get("quantity", 1))
                        if quantity == int(quantity):
                            quantity = int(quantity)
                    except (ValueError, TypeError):
                        quantity = 1
                        
                    try:
                        price = float(item.get("price", 0))
                    except (ValueError, TypeError):
                        price = 0.0
                        
                    try:
                        total = float(item.get("total", 0))
                    except (ValueError, TypeError):
                        total = 0.0
                        
                    try:
                        tax_amount = float(item.get("tax_amount", 0)) if item.get("tax_amount") else 0.0
                    except (ValueError, TypeError):
                        tax_amount = 0.0
                        
                    try:
                        final_price = float(item.get("final_price", 0)) if item.get("final_price") else total
                    except (ValueError, TypeError):
                        final_price = total
                        
                    try:
                        savings = float(item.get("savings", 0)) if item.get("savings") else None
                    except (ValueError, TypeError):
                        savings = None
                    
                    item_data = {
                        "Receipt ID": receipt_id,
                        "Date": date,
                        "Merchant": merchant,
                        "Description": item.get("description", ""),
                        "Quantity": quantity,
                        "Unit Price": price,
                        "Tax Status": item.get("tax_status", ""),
                        "Tax Amount": tax_amount,
                        "Final Price": final_price,
                        "Total": total,
                        "Category": item.get("category", ""),
                        "Savings": savings,
                        "Notes": ""
                    }
                    all_items.append(item_data)
            
            # Create pandas DataFrame and save to CSV
            if all_items:
                import pandas as pd
                df = pd.DataFrame(all_items)
                
                # Set numeric dtypes for appropriate columns
                numeric_columns = ["Quantity", "Unit Price", "Total", "Tax Amount", "Final Price", "Savings"]
                for col in numeric_columns:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce')
                
                df.to_csv(output_file_path, index=False)
                logger.info(f"Saved CSV with numeric values to {output_file_path}")
                print(f"✅ CSV file saved to {output_file_path}")
                return output_file_path
            else:
                logger.warning("No items to save to CSV")
                return None
                
        except Exception as e:
            logger.error(f"Error saving CSV: {e}")
            print(f"❌ Error saving CSV: {e}")
            return None

    # --- Helper function for applying column widths --- 
    def _apply_column_widths(self, ws, headers, data, raw_headers=None):
        """Helper to calculate and apply column widths using openpyxl."""
        if raw_headers is None:
            raw_headers = headers # Use cleaned headers if raw not provided
            
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            raw_header = raw_headers[col_idx-1] # Get corresponding raw header for data lookup
            
            # Determine if data is DataFrame or list of dicts
            if isinstance(data, pd.DataFrame):
                try:
                    # Get max length from DataFrame column (convert to string)
                    col_data = data[header].astype(str)
                    max_len_data = col_data.str.len().max()
                    if pd.notna(max_len_data):
                        max_len = max(max_len, int(max_len_data))
                except Exception as e:
                    logger.warning(f"Could not get max length from DataFrame column '{header}': {e}")
            else: # Assume list of dicts
                for record in data:
                    cell_val_str = str(record.get(raw_header) or "")
                    max_len = max(max_len, len(cell_val_str))
                    
            adjusted_width = min(max(max_len, 10) + 2, 60)
            ws.column_dimensions[col_letter].width = adjusted_width
            
    def _apply_data_validation(self, ws, column_idx):
        """Apply data validation for category dropdown lists."""
        try:
            # The method is no longer needed as we're handling validation directly 
            # in the _save_results_excel method with a simplified approach
            logger.info("Data validation now handled in main Excel save method")
        except Exception as e:
            logger.error(f"Error applying data validation: {e}")
            pass  # Continue without validation if it fails

    def _save_results_simple_excel(self, results, output_file_path):
        """Save results as a simple Excel file with minimal formatting to prevent XML corruption."""
        logger.info(f"Attempting to save simplified Excel file to: {output_file_path}")
        try:
            # Get just the items data
            _, items_data, _ = self._prepare_excel_data(results)
            
            if not items_data:
                logger.warning("No item data to save to Excel")
                return False
                
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Receipt Items"
            
            # Get and clean header names
            headers = []
            for key in items_data[0].keys():
                # Sanitize header names
                clean_header = ''.join(c if c.isalnum() or c == '_' else '_' for c in str(key))
                headers.append(clean_header)
            
            # Write header row
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
            
            # Write data rows with minimal formatting
            for row_idx, item in enumerate(items_data, 2):
                for col_idx, (header, key) in enumerate(zip(headers, item.keys()), 1):
                    # Sanitize value
                    value = item[key]
                    if value is None:
                        value = ""
                    elif isinstance(value, str):
                        # Remove control characters
                        value = ''.join(c for c in value if ord(c) >= 32 or c in '\r\n\t')
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply minimal currency formatting
                    if any(money_term in key for money_term in ['Price', 'Total', 'Amount', 'Savings']):
                        cell.number_format = '#,##0.00'
            
            # Simple column width adjustment
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                        except:
                            pass
                adjusted_width = min(max_length + 2, 40)  # Cap width at 40
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the simplified workbook
            wb.save(output_file_path)
            print(f"✅ Simplified Excel file saved to {output_file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error saving simplified Excel: {e}", exc_info=True)
            print(f"❌ Error saving simplified Excel: {e}")
            return False
    
    def save_results(self, results, output_dir="./output", include_raw=False):
        """Save analysis results to JSON and Excel files."""
        if not results:
            logger.warning("No results to save")
            return {}
        
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Generate timestamp for filenames
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Save to JSON
        json_path = os.path.join(output_dir, f"receipt_analysis_{timestamp}.json")
        try:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            logger.info(f"Saved JSON to {json_path}")
            print(f"✅ JSON file saved to {json_path}")
        except Exception as e:
            logger.error(f"Error saving JSON: {e}")
            print(f"❌ Error saving JSON: {e}")
        
        # Try to save as Excel
        excel_path = os.path.join(output_dir, f"receipt_analysis_{timestamp}.xlsx")
        
        # First try the full-featured Excel save
        excel_success = False
        try:
            excel_success = self._save_results_excel(results, excel_path)
        except Exception as e:
            logger.error(f"Error in primary Excel saving: {e}", exc_info=True)
            print(f"❌ Error in primary Excel saving: {e}")
            excel_success = False
        
        # If the full-featured Excel save fails, try the simplified version
        if not excel_success:
            try:
                simple_excel_path = os.path.join(output_dir, f"receipt_analysis_simple_{timestamp}.xlsx")
                logger.info("Attempting simplified Excel save as fallback")
                self._save_results_simple_excel(results, simple_excel_path)
            except Exception as e:
                logger.error(f"Error in simplified Excel saving: {e}", exc_info=True)
                print(f"❌ Error in simplified Excel saving: {e}")
                # As a last resort, save to CSV
                csv_path = os.path.join(output_dir, f"receipt_analysis_{timestamp}.csv")
                try:
                    self._save_results_csv(results, csv_path)
                except Exception as csv_err:
                    logger.error(f"Error creating fallback CSV: {csv_err}")
        
        return {
            "json_path": json_path,
            "excel_path": excel_path if excel_success else None
        }

    def _prepare_excel_data(self, results):
        """Helper to structure data for different Excel sheets."""
        receipt_data = []
        items_data = []
        tax_data = []

        for idx, result in enumerate(results):
            receipt_name = result.get("image_id", f"Receipt_{idx+1}")
            merchant_data = result.get("merchant", {})
            transaction_data = result.get("transaction", {})
            payment_data = result.get("payment", {})
            tax_details = transaction_data.get("tax_details", {})

            # --- Receipt Summary Data ---
            # Convert relevant fields to numeric early
            rec_subtotal = self._clean_currency(transaction_data.get("subtotal"))
            rec_total_tax = self._clean_currency(tax_details.get("total_tax"))
            rec_total_amount = self._clean_currency(transaction_data.get("total"))
            rec_payment_amount = self._clean_currency(payment_data.get("amount"))
            rec_change_due = self._clean_currency(payment_data.get("change_due"))
            
            record = {
                "Receipt ID": receipt_name,
                "Date": transaction_data.get("date", ""),
                "Time": transaction_data.get("time", ""),
                "Merchant": merchant_data.get("name", ""),
                "Merchant Address": merchant_data.get("address", ""),
                "Subtotal": rec_subtotal,
                "Total Tax": rec_total_tax,
                "Total Amount": rec_total_amount,
                "Payment Amount": rec_payment_amount,
                "Change Due": rec_change_due,
                "Notes": ""
            }
            receipt_data.append(record)

            # --- Items Detail Data ---
            for item in result.get("items", []):
                # Use the main categorization logic
                category = self._categorize_item(item.get("description", ""))
                
                # Clean numeric values
                item_unit_price = self._clean_currency(item.get("price"))
                item_total = self._clean_currency(item.get("total"))
                item_tax_amount = self._clean_currency(item.get("tax_amount"))
                item_final_price = self._clean_currency(item.get("final_price"))
                item_savings = self._clean_currency(item.get("savings"))
                
                item_record = {
                    "Receipt ID": receipt_name,
                    "Date": transaction_data.get("date", ""),
                    "Merchant": merchant_data.get("name", ""),
                    "Description": item.get("description", ""),
                    "Quantity": self._clean_quantity(item.get("quantity")), 
                    "Unit Price": item_unit_price,
                    "Total": item_total,
                    "Tax Status": item.get("tax_status", ""),
                    "Tax Amount": item_tax_amount,
                    "Final Price": item_final_price,
                    "Category": category if category != "Other" else "Uncategorized",
                    "Savings": item_savings,
                    "Notes": ""
                }
                items_data.append(item_record)

        return receipt_data, items_data, tax_data

def main():
    """
    Main function to process receipt files and generate structured output
    """
    try:
        # Set up output directory
        output_dir = os.environ.get('OUTPUT_DIR', 'output')
        os.makedirs(output_dir, exist_ok=True)
        
        print("\n===== RECEIPT DIGITIZER FOR BUDGETING =====")
        print("This tool will process receipts and generate structured data for budgeting.")
        print("Outputs will be saved as both JSON and Excel files.")
        
        # Check for required packages
        missing_packages = []
        try:
            import pandas
        except ImportError:
            missing_packages.append("pandas")
        
        try:
            import openpyxl
        except ImportError:
            missing_packages.append("openpyxl")
            
        if missing_packages:
            print("\n⚠️ WARNING: Missing required packages for full functionality")
            print(f"The following packages are required but not installed: {', '.join(missing_packages)}")
            print("\nPlease install these packages before running the script:")
            installation_cmds = [f"pip install {pkg}" for pkg in missing_packages]
            print("\n".join(installation_cmds))
            print("\nJSON output will still be generated, but structured Excel output requires these packages.")
            
            user_input = input("\nDo you want to continue with limited functionality? (y/n): ")
            if user_input.lower() != 'y':
                print("Exiting script. Please install the required packages and run again.")
                return
        
        # Check required environment variables
        required_vars = ['VISION_ENDPOINT', 'VISION_API_KEY']
        missing_vars = [var for var in required_vars if not os.environ.get(var)]
        
        if missing_vars:
            print("\n❌ ERROR: Missing required environment variables")
            print(f"The following variables are required: {', '.join(missing_vars)}")
            print("\nPlease set these environment variables before running the script:")
            print("  - VISION_ENDPOINT: Your Azure Document Intelligence endpoint")
            print("  - VISION_API_KEY: Your Azure Document Intelligence API key")
            return
            
        # Initialize the receipt analyzer
        analyzer = AzureReceiptAnalyzer(
            os.environ['VISION_ENDPOINT'],
            os.environ['VISION_API_KEY']
        )
        
        # Define test scenarios
        test_scenarios = [
            {
                "name": "Local Receipt Processing",
                "files": ["receipts/*.jpg"],
                "max_receipts": 10
            },
            {
                "name": "Directory Receipt Processing",
                "files": ["path/to/receipts"],
                "max_receipts": 5
            },
            {
                "name": "Pattern Receipt Processing",
                "files": ["path/to/receipts/*.jpg"],
                "max_receipts": 5
            }
        ]
        
        # Process each test scenario
        for scenario in test_scenarios:
            print(f"\n===== {scenario['name'].upper()} =====")
            for pattern in scenario['files']:
                file_paths = glob.glob(pattern)
                if not file_paths:
                    print(f"No files found matching pattern: {pattern}")
                    continue
                    
                print(f"\n📄 Found {len(file_paths)} receipt files to process:")
                for path in file_paths[:5]:  # Show first 5 files
                    print(f"  - {os.path.basename(path)}")
                if len(file_paths) > 5:
                    print(f"  ... and {len(file_paths) - 5} more")
                
                results = analyzer.process_receipts(file_paths, max_receipts=scenario['max_receipts'])
                
                # Generate summary report
                if results:
                    summary_report = analyzer.generate_summary_report(results)
                    print(f"\n📊 Results for {scenario['name']}:")
                    print(f"  • Total Spending: ${summary_report['total_spending']:.2f}")
                    print(f"  • Receipts Processed: {summary_report['receipt_count']}")
                    print(f"  • Unique Merchants: {summary_report['merchant_count']}")
                    print(f"  • Total Items: {summary_report['item_count']}")
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        logger.error(f"Error in main function: {str(e)}")
        logger.error(traceback.format_exc())
        print("\nFor detailed error information, check the logs.")

if __name__ == "__main__":
    main()